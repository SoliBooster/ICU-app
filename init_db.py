# ================= 数据库初始化（从 Excel 导入） =================
"""
从 ICU数据.xlsx 读取 8 个分表，自动创建 SQL 表并导入数据。
支持 SQLite（本地开发）和 MySQL（生产部署）。
"""
import os
import re
import math
import pymysql
import openpyxl
from werkzeug.security import generate_password_hash
from config import DB_PATH, EXCEL_PATH, MYSQL_HOST, MYSQL_PORT, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DB
from db import get_db, use_mysql, placeholder


def sanitize(name: str) -> str:
    """把任意字符串转为合法的 SQL 标识符"""
    s = re.sub(r'[^a-zA-Z0-9_]', '_', str(name))
    s = re.sub(r'_+', '_', s).strip('_')
    if not s:
        s = 't'
    if s[0].isdigit():
        s = 't_' + s
    return s[:50]


def parse_float(v):
    """把单元格值转为 float，无法解析返回 None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s2 = re.sub(r'[<≥≤≈>~]', '', s)
    try:
        return float(s2)
    except ValueError:
        return None


def col_type():
    """列数据类型：MySQL 用 DOUBLE，SQLite 用 REAL"""
    return 'DOUBLE' if use_mysql() else 'REAL'


def text_type():
    """文本类型：MySQL 用 VARCHAR，SQLite 用 TEXT"""
    return 'VARCHAR(255)' if use_mysql() else 'TEXT'


def init_db(force=False):
    """读取 Excel，创建表并导入数据"""
    if force and not use_mysql() and os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    if use_mysql():
        # MySQL 模式：先确保数据库存在
        conn = pymysql.connect(
            host=MYSQL_HOST, port=MYSQL_PORT,
            user=MYSQL_USER, password=MYSQL_PASSWORD,
            charset="utf8mb4"
        )
        try:
            with conn.cursor() as cur:
                cur.execute(f"CREATE DATABASE IF NOT EXISTS `{MYSQL_DB}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
            conn.commit()
        finally:
            conn.close()



    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = get_db()
    cur = conn.cursor()

    # 1. 创建元数据表
    p = placeholder()
    if use_mysql():
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS _meta_columns (
                id INT AUTO_INCREMENT PRIMARY KEY,
                sheet_key VARCHAR(50) NOT NULL,
                sheet_display VARCHAR(200) NOT NULL,
                col_key VARCHAR(50) NOT NULL,
                col_display VARCHAR(200) NOT NULL,
                col_order INT NOT NULL,
                sheet_order INT NOT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    else:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS _meta_columns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_key TEXT NOT NULL,
                sheet_display TEXT NOT NULL,
                col_key TEXT NOT NULL,
                col_display TEXT NOT NULL,
                col_order INTEGER NOT NULL,
                sheet_order INTEGER NOT NULL
            )
        """)

    sheets_info = []

    for order, sname in enumerate(wb.sheetnames):
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = rows[0]
        col_displays = [c for c in header[1:] if c is not None and str(c).strip() != '']
        if not col_displays:
            continue

        sheet_key = 's_' + str(order + 1)
        tname = 't_' + sanitize(sname)[:40] + '_' + str(order + 1)

        ct = col_type()
        tt = text_type()
        auto = 'AUTO_INCREMENT' if use_mysql() else 'AUTOINCREMENT'

        # 2. 创建动态 SQL 表
        col_defs = ', '.join([f'col_{i + 1} {ct}' for i in range(len(col_displays))])
        pid_col = 'patient_id INT DEFAULT 1' if use_mysql() else 'patient_id INTEGER DEFAULT 1'
        if use_mysql():
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS `{tname}` (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    patient_id INT DEFAULT 1,
                    row_label {tt} DEFAULT '',
                    {col_defs}
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        else:
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {tname} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    patient_id INTEGER DEFAULT 1,
                    row_label TEXT DEFAULT '',
                    {col_defs}
                )
            """)

        sheets_info.append({
            'key': sheet_key,
            'display': sname,
            'table_name': tname,
            'columns': [{'key': f'col_{i + 1}', 'display': str(col_displays[i])} for i in range(len(col_displays))],
            'order': order,
        })

        # 3. 写入元数据
        cur.execute(f"DELETE FROM _meta_columns WHERE sheet_key={p}", (sheet_key,))
        for ci, col in enumerate(col_displays):
            cur.execute(
                f"INSERT INTO _meta_columns(sheet_key, sheet_display, col_key, col_display, col_order, sheet_order) VALUES({p},{p},{p},{p},{p},{p})",
                (sheet_key, sname, f'col_{ci + 1}', str(col), ci, order)
            )

        # 4. 导入数据
        cur.execute(f"DELETE FROM {tname}")
        data_rows = rows[1:]
        for r in data_rows:
            if r[0] is None and all(v is None for v in r[1:]):
                continue
            row_label = str(r[0]) if r[0] is not None else ''
            values = [row_label]
            placeholders = [p]
            col_names = ['row_label']
            for i in range(len(col_displays)):
                v = r[i + 1] if (i + 1) < len(r) else None
                values.append(parse_float(v))
                placeholders.append(p)
                col_names.append(f'col_{i + 1}')
            cur.execute(f"INSERT INTO {tname}({', '.join(col_names)}) VALUES({','.join(placeholders)})", values)

    # 5. 创建用户表
    if use_mysql():
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                role VARCHAR(20) NOT NULL DEFAULT 'user',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    else:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

    # 6. 预置管理员账号
    existing = cur.execute(f"SELECT id FROM users WHERE username={p}", ('admin',))
    row = cur.fetchone() if use_mysql() else existing.fetchone()
    if use_mysql():
        existing_row = cur.fetchone()
        if not existing_row:
            cur.execute(
                f"INSERT INTO users(username, password_hash, role) VALUES({p},{p},{p})",
                ('admin', generate_password_hash('admICU123456'), 'admin')
            )
            print('已创建管理员账号：admin')
    else:
        if not row:
            cur.execute(
                f"INSERT INTO users(username, password_hash, role) VALUES({p},{p},{p})",
                ('admin', generate_password_hash('admICU123456'), 'admin')
            )
            print('已创建管理员账号：admin')

    # 7. 创建病人表
    if use_mysql():
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS patients (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                view_password_hash VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    else:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS patients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                view_password_hash TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

    # 8. 预置默认病人
    if use_mysql():
        cur.execute(f"SELECT id FROM patients WHERE name={p}", ('杜桂琴',))
        patient_row = cur.fetchone()
    else:
        cur.execute(f"SELECT id FROM patients WHERE name={p}", ('杜桂琴',))
        patient_row = cur.fetchone()

    if not patient_row:
        cur.execute(
            f"INSERT INTO patients(name, view_password_hash) VALUES({p},{p})",
            ('杜桂琴', generate_password_hash('123456'))
        )
        print('已创建默认患者：杜桂琴（查看密码：123456）')

    conn.commit()
    conn.close()
    wb.close()

    print(f"数据库初始化完成：{len(sheets_info)} 个分表")
    for s in sheets_info:
        print(f"  {s['key']}: {s['display']} ({len(s['columns'])} 列)")


if __name__ == '__main__':
    init_db(force=True)
